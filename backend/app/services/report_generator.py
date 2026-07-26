import io
from datetime import datetime
from typing import List, Any, Optional, Dict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape, A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle


def export_to_excel(
    title: str,
    headers: List[str],
    rows: List[List[Any]],
    summary_data: Optional[Dict[str, Any]] = None
) -> bytes:
    """
    Generate file Excel (.xlsx) bergaya profesional dan rapi menggunakan openpyxl (B-24).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan"

    # Styling definitions
    font_title = Font(name="Calibri", size=16, bold=True, color="0A5C36")
    font_sub = Font(name="Calibri", size=10, italic=True, color="555555")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_data = Font(name="Calibri", size=11)
    font_bold = Font(name="Calibri", size=11, bold=True)
    
    fill_header = PatternFill(start_color="0A5C36", end_color="0A5C36", fill_type="solid")
    fill_zebra = PatternFill(start_color="F9F8F5", end_color="F9F8F5", fill_type="solid")
    
    border_thin = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    # 1. Title & Timestamp
    ws["A1"] = title.upper()
    ws["A1"].font = font_title
    ws["A2"] = f"Generated at: {datetime.utcnow().strftime('%d-%m-%Y %H:%M:%S UTC')} | SMK PTDARRAHMAN"
    ws["A2"].font = font_sub
    
    current_row = 4

    # 2. Summary Data Section (if provided)
    if summary_data:
        ws.cell(row=current_row, column=1, value="RINGKASAN LAPORAN").font = Font(name="Calibri", size=12, bold=True, color="0A5C36")
        current_row += 1
        for k, v in summary_data.items():
            ws.cell(row=current_row, column=1, value=k).font = font_bold
            val_cell = ws.cell(row=current_row, column=2, value=v)
            val_cell.font = font_data
            if isinstance(v, (int, float)) and "Rp" in str(k) or "Total" in str(k) or "Nominal" in str(k):
                val_cell.number_format = '#,##0'
            current_row += 1
        current_row += 1

    # 3. Table Headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=current_row, column=col_idx, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center" if col_idx == 1 or "No" in header else "left", vertical="center")
        cell.border = border_thin
    
    ws.row_dimensions[current_row].height = 24
    current_row += 1

    # 4. Table Rows
    for row_idx, r_data in enumerate(rows):
        is_even = (row_idx % 2 == 0)
        for col_idx, val in enumerate(r_data, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.font = font_data
            cell.border = border_thin
            if not is_even:
                cell.fill = fill_zebra
            
            # Auto format numbers/currency
            if isinstance(val, (int, float)):
                if "Rp" in headers[col_idx-1] or "Nominal" in headers[col_idx-1] or "Total" in headers[col_idx-1] or "Infaq" in headers[col_idx-1] or "Bayar" in headers[col_idx-1]:
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal="center")
            elif isinstance(val, str) and (val.startswith("Rp") or val.replace(".", "").isdigit()):
                cell.alignment = Alignment(horizontal="right" if val.startswith("Rp") else "left")
            else:
                cell.alignment = Alignment(horizontal="center" if col_idx == 1 else "left")
                
        ws.row_dimensions[current_row].height = 20
        current_row += 1

    # 5. Auto-fit column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row < 3: # skip title banner
                continue
            val_str = str(cell.value or "")
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    out_buf = io.BytesIO()
    wb.save(out_buf)
    excel_bytes = out_buf.getvalue()
    out_buf.close()
    return excel_bytes


def export_to_pdf(
    title: str,
    subtitle: str,
    headers: List[str],
    rows: List[List[Any]],
    summary_data: Optional[Dict[str, Any]] = None,
    is_landscape: bool = False
) -> bytes:
    """
    Generate file PDF laporan formal menggunakan ReportLab (B-24).
    """
    buffer = io.BytesIO()
    page_size = landscape(A4) if (is_landscape or len(headers) > 6) else A4
    doc = SimpleDocTemplate(
        buffer,
        pagesize=page_size,
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'RepTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=16,
        textColor=colors.HexColor("#0A5C36"),
        alignment=1,
        spaceAfter=6
    )
    sub_style = ParagraphStyle(
        'RepSub',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        textColor=colors.HexColor("#555555"),
        alignment=1,
        spaceAfter=15
    )
    normal_cell = ParagraphStyle('CellNorm', parent=styles['Normal'], fontName='Helvetica', fontSize=9, leading=11)
    header_cell = ParagraphStyle('CellHead', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=9, leading=11, textColor=colors.white, alignment=1)

    elements = []

    # Title & Subtitle
    elements.append(Paragraph(title.upper(), title_style))
    elements.append(Paragraph(f"{subtitle} | SMK PTDARRAHMAN | Dicetak: {datetime.utcnow().strftime('%d-%m-%Y %H:%M')}", sub_style))
    elements.append(Spacer(1, 10))

    # Summary Section
    if summary_data:
        sum_rows = []
        for k, v in summary_data.items():
            val_str = f"Rp {v:,.2f}" if isinstance(v, (int, float)) and ("Total" in k or "Nominal" in k or "Rp" in k or "Infaq" in k or "Target" in k or "Terkumpul" in k or "Sisa" in k or "Bayar" in k or "Saldo" in k or "Hutang" in k) else str(v)
            sum_rows.append([Paragraph(f"<b>{k}</b>", normal_cell), Paragraph(val_str, normal_cell)])
        
        t_sum = Table(sum_rows, colWidths=[200, 300])
        t_sum.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor("#F9F8F5")),
            ('BOX', (0,0), (-1,-1), 0.5, colors.HexColor("#CCCCCC")),
            ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor("#EEEEEE")),
            ('TOPPADDING', (0,0), (-1,-1), 5),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ('LEFTPADDING', (0,0), (-1,-1), 8),
        ]))
        elements.append(t_sum)
        elements.append(Spacer(1, 15))

    # Table Content
    table_data = [[Paragraph(h, header_cell) for h in headers]]
    for r in rows:
        row_cells = []
        for idx, val in enumerate(r):
            if isinstance(val, (int, float)):
                if "Rp" in headers[idx] or "Nominal" in headers[idx] or "Total" in headers[idx] or "Infaq" in headers[idx] or "Bayar" in headers[idx] or "Target" in headers[idx]:
                    val_str = f"Rp {val:,.2f}"
                else:
                    val_str = f"{val:,.0f}"
            else:
                val_str = str(val or "-")
            
            align = 2 if isinstance(val, (int, float)) or "Rp" in str(val) else 1 if idx == 0 else 0
            cell_style = ParagraphStyle(f'C_{idx}', parent=normal_cell, alignment=align)
            row_cells.append(Paragraph(val_str, cell_style))
        table_data.append(row_cells)

    # Calculate proportional column widths
    avail_width = page_size[0] - 60
    num_cols = len(headers)
    col_widths = [avail_width / num_cols] * num_cols
    # Adjust first col (No / ID) if possible
    if num_cols >= 3 and ("No" in headers[0] or "ID" in headers[0]):
        col_widths[0] = 40
        rem_width = avail_width - 40
        for i in range(1, num_cols):
            col_widths[i] = rem_width / (num_cols - 1)

    t_data = Table(table_data, colWidths=col_widths, repeatRows=1)
    t_data.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#0A5C36")),
        ('ALIGN', (0,0), (-1,0), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#DDDDDD")),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('LEFTPADDING', (0,0), (-1,-1), 6),
        ('RIGHTPADDING', (0,0), (-1,-1), 6),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor("#FDFBF7")]),
    ]))
    elements.append(t_data)

    doc.build(elements)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    return pdf_bytes
